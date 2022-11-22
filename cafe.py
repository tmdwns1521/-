# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'cafe.ui'
#
# Created by: PyQt5 UI code generator 5.14.1
#
# WARNING! All changes made in this file will be lost!


from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import time, os

dir = (os.path.dirname(os.path.realpath(__file__)))

class Thread2(QThread):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent

    def run(self):
        f = open(f"{dir}/id_list.txt","r")
        f = f.read().split("\n")
        dict = {}
        for i in f:
            dict[(i.split("\t")[0])] = []
        for i in f:
            if i == "":
                pass
            else:
                phonenum = (i.split("\t")[1])
                writetime = (i.split("\t")[2])
                dict[(i.split("\t")[0])].append([writetime,phonenum])
        lists = []
        for key,value in dict.items():
            try:
                values = sorted(value,reverse=True)
                if not values[0][0]:
                    pass
                else:
                    phonenums = (values[0][0])
                    writetime = (values[0][1])
                    lists.append(key+"\t"+phonenums+"\t"+writetime)
            except Exception as ex:
                pass
        fws = open(f"{dir}/id_list.txt","w+")
        for i in lists:
            name = i.split("\t")[0]
            write_time = i.split("\t")[2]
            phone_num = i.split("\t")[1]
            fws.write(name+"\t"+write_time+"\t"+phone_num+"\n")

class Thread1(QThread):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent

    def run(self):
        ids = self.parent.lineEdit.text()
        pwss = self.parent.lineEdit_2.text()
        # c#!/usr/bin/env python
        # coding: utf-8
        from threading import Thread
        import requests
        import json,time
        from multiprocessing import Process
        from bs4 import BeautifulSoup
        import urllib.parse,os,datetime,openpyxl
        from selenium.webdriver.common.keys import Keys
        from openpyxl import load_workbook
        from urllib.parse import urljoin
        from selenium import webdriver
        import re
        import uuid
        import requests
        import rsa
        import lzstring
        import queue
        from urllib3.util.retry import Retry
        from requests.adapters import HTTPAdapter

        dir = (os.path.dirname(os.path.realpath(__file__)))
        now = datetime.datetime.now()
        nowDatetime = now.strftime('%Y-%m-%d_%H_%M_%S')

        def encrypt(key_str, uid, upw):
            def naver_style_join(l):
                return ''.join([chr(len(s)) + s for s in l])

            sessionkey, keyname, e_str, n_str = key_str.split(',')
            e, n = int(e_str, 16), int(n_str, 16)

            message = naver_style_join([sessionkey, uid, upw]).encode()

            pubkey = rsa.PublicKey(e, n)
            encrypted = rsa.encrypt(message, pubkey)

            return keyname, encrypted.hex()


        def encrypt_account(uid, upw):
            key_str = requests.get('https://nid.naver.com/login/ext/keys.nhn').content.decode("utf-8")
            return encrypt(key_str, uid, upw)


        def naver_session(nid, npw):
            encnm, encpw = encrypt_account(nid, npw)

            s = requests.Session()
            retries = Retry(
                total=5,
                backoff_factor=0.1,
                status_forcelist=[500, 502, 503, 504]
            )
            s.mount('https://', HTTPAdapter(max_retries=retries))
            request_headers = {
                'User-agent': 'Mozilla/5.0'
            }

            bvsd_uuid = uuid.uuid4()
            encData = '{"a":"%s-4","b":"1.3.4","d":[{"i":"id","b":{"a":["0,%s"]},"d":"%s","e":false,"f":false},{"i":"%s","e":true,"f":false}],"h":"1f","i":{"a":"Mozilla/5.0"}}' % (bvsd_uuid, nid, nid, npw)
            bvsd = '{"uuid":"%s","encData":"%s"}' % (bvsd_uuid, lzstring.LZString.compressToEncodedURIComponent(encData))

            resp = s.post('https://nid.naver.com/nidlogin.login', data={
                'svctype': '0',
                'enctp': '1',
                'encnm': encnm,
                'enc_url': 'http0X0.0000000000001P-10220.0000000.000000www.naver.com',
                'url': 'www.naver.com',
                'smart_level': '1',
                'encpw': encpw,
                'bvsd': bvsd
            }, headers=request_headers)

            finalize_url = re.search(r'location\.replace\("([^"]+)"\)', resp.content.decode("utf-8")).group(1)
            s.get(finalize_url)

            return s

        session = naver_session(f'{ids}', f'{pwss}')

        # def func1(ids):
        headers = {
        'content-type': 'application/json;charset=utf-8',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36'
        }
        wb=load_workbook(filename=f'{dir}/아이디체크.xlsx',read_only=False,data_only=False)
        ws=wb['Sheet1']
        time.sleep(1.5)
        while True:
            if ws['A1'].value == None:
                break
            check_ids = ws['A1'].value
            self.parent.textBrowser.append(str(check_ids))
            time.sleep(0.7)
            number_check = False
            url = f'https://apis.naver.com/cafe-web/cafe-mobile/CafeMemberProfile?cafeId=10050146&memberKey=&memberId={check_ids}&requestFrom=A'
            r = session.get(url, headers = headers).text
            statuss = json.loads(r)['message']['status']
            if statuss == '500':
                print(statuss)
                pass
            else:
                r_json = json.loads(r)['message']['result']['memberKey']
                time.sleep(0.7)
                url = f'https://apis.naver.com/cafe-web/cafe-mobile/CafeMemberNetworkArticleList?search.cafeId=10050146&search.memberKey={r_json}&search.perPage=75&search.page=1&requestFrom=A'
                r = session.get(url, headers = headers).text
                r_json = json.loads(r)['message']['result']['articleList']
                self.parent.textBrowser.append(f"{check_ids} 글 발행 갯수 : {len(r_json)}")
                if not r_json:
                    pass
                counts = len(r_json)
                def numbers_check(rj, check_ids):
                    numbers = ''
                    time.sleep(0.9)
                    articleid = rj['articleid']
                    clubid = rj['clubid']
                    urls = f'https://apis.naver.com/cafe-web/cafe-articleapi/v2/cafes/{clubid}/articles/{articleid}?query=&useCafeId=true&requestFrom=A'
                    r = session.get(urls, headers = headers).text
                    if "비정상적" in r:
                        time.sleep(600)
                    try:
                        new_text_level = json.loads(r)['result']["errorCode"]
                    except Exception as ex:
                        print(ex)
                        new_date = json.loads(r)['result']['article']['writeDate']
                        new_text = json.loads(r)['result']['article']['contentHtml']
                        new_text = new_text.replace(".","").replace("-","")
                        new_text = BeautifulSoup(new_text,'html.parser').text
                        print(new_text)
                        p = re.compile('010-?[0-9]{4}-?[0-9]{4}')
                        result = p.findall(new_text)
                        if not result:
                            p = re.compile('010 ?[0-9]{4} ?[0-9]{4}')
                            result = p.findall(new_text)
                            if not result:
                                new_text = new_text.replace(" ","")
                                number_reset1 = [["공","0"],["일","1"],["이","2"],["삼","3"],["사","4"],["오","5"],["육","6"],["칠","7"],["팔","8"],["구","9"]]
                                number_reset2 = [['ㅇ',"0"],["영","0"],['i',"1"],['L',"1"],['l',"1"],['|',"1"],['투',"2"],['석',"3"],['셋',"3"],['쓰리',"3"],['포',"4"],['q',"9"],['빵',"0"],['o',"0"],['O',"0"],['@',"0"],["하나","1"],["둘","2"],["셋","3"],["넷","4"],["다섯","5"],["여섯","6"],["일곱","7"],["여덟","8"],["아홉","9"]]
                                for i in number_reset1:
                                    new_text = new_text.replace(i[0],i[1])
                                for i in number_reset2:
                                    new_text = new_text.replace(i[0],i[1])
                                p = re.compile('010-?[0-9]{4}-?[0-9]{4}')
                                result = p.findall(new_text)
                                if not result:
                                    p = re.compile('010 ?[0-9]{4} ?[0-9]{4}')
                                    result = p.findall(new_text)
                                    if not result:
                                        pass
                                    else:
                                        numbers = result[0]
                                else:
                                    numbers = result[0]
                            else:
                                numbers = result[0]
                        else:
                            numbers = result[0]
                        print(numbers)
                        print("-----------------------------------------------------")
                    if numbers != '':
                        main_id = str(check_ids) + "\t" + str(numbers) + "\t" + str(new_date)
                        f = open(f"{dir}/id_list.txt","a+")
                        f.write(main_id + "\n")
                        self.parent.textBrowser.append(main_id)
                        number_check = True
                for ii,rj in enumerate(r_json):
                    thread1 = Thread(target=numbers_check, args=(rj,check_ids))
                    thread1.start()
            if number_check == False:
                time.sleep(0.7)
                url = f'https://apis.naver.com/cafe-web/cafe-mobile/CafeMemberProfile?cafeId=14940923&memberKey=&memberId={check_ids}&requestFrom=A'
                r = session.get(url, headers = headers).text
                statuss = json.loads(r)['message']['status']
                if statuss == '500':
                    print(statuss)
                    pass
                else:
                    r_json = json.loads(r)['message']['result']['memberKey']
                    time.sleep(0.7)
                    url = f'https://apis.naver.com/cafe-web/cafe-mobile/CafeMemberNetworkArticleList?search.cafeId=14940923&search.memberKey={r_json}&search.perPage=75&search.page=1&requestFrom=A'
                    r = session.get(url, headers = headers).text
                    r_json = json.loads(r)['message']['result']['articleList']
                    self.parent.textBrowser.append(f"{check_ids} 글 발행 갯수 : {len(r_json)}")
                    if not r_json:
                        pass
                    counts = len(r_json)
                    def numbers_check(rj, check_ids):
                        numbers = ''
                        time.sleep(0.9)
                        articleid = rj['articleid']
                        clubid = rj['clubid']
                        urls = f'https://apis.naver.com/cafe-web/cafe-articleapi/v2/cafes/{clubid}/articles/{articleid}?query=&useCafeId=true&requestFrom=A'
                        r = session.get(urls, headers = headers).text
                        if "비정상적" in r:
                            time.sleep(600)
                        try:
                            new_text_level = json.loads(r)['result']["errorCode"]
                            print(new_text_level)
                        except Exception as ex:
                            print(ex)
                            new_date = json.loads(r)['result']['article']['writeDate']
                            new_text = json.loads(r)['result']['article']['contentHtml']
                            new_text = new_text.replace(".","").replace("-","")
                            new_text = BeautifulSoup(new_text,'html.parser').text
                            p = re.compile('010-?[0-9]{4}-?[0-9]{4}')
                            result = p.findall(new_text)
                            if not result:
                                p = re.compile('010 ?[0-9]{4} ?[0-9]{4}')
                                result = p.findall(new_text)
                                if not result:
                                    new_text = new_text.replace(" ","")
                                    number_reset1 = [["공","0"],["일","1"],["이","2"],["삼","3"],["사","4"],["오","5"],["육","6"],["칠","7"],["팔","8"],["구","9"]]
                                    number_reset2 = [['ㅇ',"0"],["영","0"],['i',"1"],['L',"1"],['l',"1"],['|',"1"],['투',"2"],['석',"3"],['셋',"3"],['쓰리',"3"],['포',"4"],['q',"9"],['빵',"0"],['o',"0"],['O',"0"],['@',"0"],["하나","1"],["둘","2"],["셋","3"],["넷","4"],["다섯","5"],["여섯","6"],["일곱","7"],["여덟","8"],["아홉","9"]]
                                    for i in number_reset1:
                                        new_text = new_text.replace(i[0],i[1])
                                    for i in number_reset2:
                                        new_text = new_text.replace(i[0],i[1])
                                    p = re.compile('010-?[0-9]{4}-?[0-9]{4}')
                                    result = p.findall(new_text)
                                    if not result:
                                        p = re.compile('010 ?[0-9]{4} ?[0-9]{4}')
                                        result = p.findall(new_text)
                                        if not result:
                                            pass
                                        else:
                                            numbers = result[0]
                                    else:
                                        numbers = result[0]
                                else:
                                    numbers = result[0]
                            else:
                                numbers = result[0]
                            print(numbers)
                            print("-----------------------------------------------------")
                        if numbers != '':
                            main_id = str(check_ids) + "\t" + str(numbers) + "\t" + str(new_date)
                            f = open(f"{dir}/id_list.txt","a+")
                            f.write(main_id + "\n")
                            self.parent.textBrowser.append(main_id)
                    for ii,rj in enumerate(r_json):
                        thread1 = Thread(target=numbers_check, args=(rj,check_ids))
                        thread1.start()
            if number_check == False:
                time.sleep(0.7)
                url = f'https://apis.naver.com/cafe-web/cafe-mobile/CafeMemberProfile?cafeId=20486145&memberKey=&memberId={check_ids}&requestFrom=A'
                r = session.get(url, headers = headers).text
                statuss = json.loads(r)['message']['status']
                if statuss == '500':
                    print(statuss)
                    pass
                else:
                    r_json = json.loads(r)['message']['result']['memberKey']
                    time.sleep(0.7)
                    url = f'https://apis.naver.com/cafe-web/cafe-mobile/CafeMemberNetworkArticleList?search.cafeId=14940923&search.memberKey={r_json}&search.perPage=75&search.page=1&requestFrom=A'
                    r = session.get(url, headers = headers).text
                    r_json = json.loads(r)['message']['result']['articleList']
                    self.parent.textBrowser.append(f"{check_ids} 글 발행 갯수 : {len(r_json)}")
                    if not r_json:
                        pass
                    counts = len(r_json)
                    def numbers_check(rj, check_ids):
                        numbers = ''
                        time.sleep(0.9)
                        articleid = rj['articleid']
                        clubid = rj['clubid']
                        urls = f'https://apis.naver.com/cafe-web/cafe-articleapi/v2/cafes/{clubid}/articles/{articleid}?query=&useCafeId=true&requestFrom=A'
                        r = session.get(urls, headers = headers).text
                        if "비정상적" in r:
                            time.sleep(600)
                        try:
                            new_text_level = json.loads(r)['result']["errorCode"]
                            print(new_text_level)
                        except Exception as ex:
                            print(ex)
                            new_date = json.loads(r)['result']['article']['writeDate']
                            new_text = json.loads(r)['result']['article']['contentHtml']
                            new_text = new_text.replace(".","").replace("-","")
                            new_text = BeautifulSoup(new_text,'html.parser').text
                            p = re.compile('010-?[0-9]{4}-?[0-9]{4}')
                            result = p.findall(new_text)
                            if not result:
                                p = re.compile('010 ?[0-9]{4} ?[0-9]{4}')
                                result = p.findall(new_text)
                                if not result:
                                    new_text = new_text.replace(" ","")
                                    number_reset1 = [["공","0"],["일","1"],["이","2"],["삼","3"],["사","4"],["오","5"],["육","6"],["칠","7"],["팔","8"],["구","9"]]
                                    number_reset2 = [['ㅇ',"0"],["영","0"],['i',"1"],['L',"1"],['l',"1"],['|',"1"],['투',"2"],['석',"3"],['셋',"3"],['쓰리',"3"],['포',"4"],['q',"9"],['빵',"0"],['o',"0"],['O',"0"],['@',"0"],["하나","1"],["둘","2"],["셋","3"],["넷","4"],["다섯","5"],["여섯","6"],["일곱","7"],["여덟","8"],["아홉","9"]]
                                    for i in number_reset1:
                                        new_text = new_text.replace(i[0],i[1])
                                    for i in number_reset2:
                                        new_text = new_text.replace(i[0],i[1])
                                    p = re.compile('010-?[0-9]{4}-?[0-9]{4}')
                                    result = p.findall(new_text)
                                    if not result:
                                        p = re.compile('010 ?[0-9]{4} ?[0-9]{4}')
                                        result = p.findall(new_text)
                                        if not result:
                                            pass
                                        else:
                                            numbers = result[0]
                                    else:
                                        numbers = result[0]
                                else:
                                    numbers = result[0]
                            else:
                                numbers = result[0]
                            print(numbers)
                            print("-----------------------------------------------------")
                        if numbers != '':
                            main_id = str(check_ids) + "\t" + str(numbers) + "\t" + str(new_date)
                            f = open(f"{dir}/id_list.txt","a+")
                            f.write(main_id + "\n")
                            self.parent.textBrowser.append(main_id)
                    for ii,rj in enumerate(r_json):
                        thread1 = Thread(target=numbers_check, args=(rj,check_ids))
                        thread1.start()
            ws.delete_rows(1)
            wb.save(f'{dir}/아이디체크.xlsx')

class Ui_Form(QWidget):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(542, 868)
        self.textBrowser = QtWidgets.QTextBrowser(Form)
        self.textBrowser.setGeometry(QtCore.QRect(20, 70, 501, 731))
        self.textBrowser.setObjectName("textBrowser")
        self.label = QtWidgets.QLabel(Form)
        self.label.setGeometry(QtCore.QRect(20, 11, 71, 51))
        self.label.setObjectName("label")
        self.pushButton = QtWidgets.QPushButton(Form)
        self.pushButton.setGeometry(QtCore.QRect(20, 812, 250, 41))
        self.pushButton.setObjectName("pushButton")
        self.pushButton_2 = QtWidgets.QPushButton(Form)
        self.pushButton_2.setGeometry(QtCore.QRect(280, 812, 240, 41))
        self.pushButton_2.setObjectName("pushButton_2")
        self.lineEdit = QtWidgets.QLineEdit(Form)
        self.lineEdit.setGeometry(QtCore.QRect(90, 20, 113, 31))
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit_2 = QtWidgets.QLineEdit(Form)
        self.lineEdit_2.setGeometry(QtCore.QRect(400, 20, 113, 31))
        self.lineEdit_2.setEchoMode(QLineEdit.Password)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.label_2 = QtWidgets.QLabel(Form)
        self.label_2.setGeometry(QtCore.QRect(310, 10, 71, 51))
        self.label_2.setObjectName("label_2")
        self.pushButton.clicked.connect(self.cafe_check)
        self.pushButton_2.clicked.connect(self.pause)

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def cafe_check(self):
        x = Thread1(self)
        x.start()
    def pause(self):
        x2 = Thread2(self)
        x2.start()

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.label.setText(_translate("Form", "네이버 ID"))
        self.pushButton.setText(_translate("Form", "작업 시작"))
        self.pushButton_2.setText(_translate("Form", "중복 제거"))
        self.label_2.setText(_translate("Form", "네이버 PW"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Form = QtWidgets.QWidget()
    ui = Ui_Form()
    ui.setupUi(Form)
    Form.show()
    sys.exit(app.exec_())
